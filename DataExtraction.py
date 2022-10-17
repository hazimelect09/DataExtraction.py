from openpyxl import load_workbook
from Framework import framewk
import datetime
import time
print('Enter starting point')
stp = int(input())
print('Enter Ending point')
end = int(input())
print('Enter the location of the chrome driver ')
ChromePATH = input()
#ChromePATH="D:\MaharahDataExtraction\chromedriver.exe"
print('Enter the Excel file location file ')
#path = "D:\MaharahDataExtraction\DataRead.xlsx"
path = input()
try:
  wb = load_workbook(path)
  sh = wb['Test']
except Exception as e:
  print('The Excel file is not found',e)
  time.sleep(4)
Exlpnt = 1
try:
    frame = framewk(ChromePATH)
except Exception as E:
    print('Check chrome driver ', E)
print('Processing....')
start = time.time()
for i in range(stp,end+1):
      print('processing page number',i)
      text = "processing page number" + str(i) + '\n'
      for j in range(10):
          try:
            frame.goto_Page(i)
            Exlpnt=Exlpnt+1
            time.sleep(1)
            sh.cell(Exlpnt,16).value=frame.open_request(j)
            time.sleep(1)
            SocialnumInsurance=frame.getSocIns()
            print(SocialnumInsurance)
            sh.cell(Exlpnt,1).value=SocialnumInsurance
            Status=frame.getStatus()
            print(Status)
            sh.cell(Exlpnt,2).value=Status
            Nationality = frame.getNationality()
            print(Nationality)
            sh.cell(Exlpnt,3).value=Nationality
            NationalID = frame.getNationalID(Nationality)
            print(NationalID)
            sh.cell(Exlpnt,4).value=NationalID
            Gender = frame.getGender()
            print(Gender)
            sh.cell(Exlpnt,5).value=Gender
            BorderNum=frame.getBordNum(Nationality)
            sh.cell(Exlpnt,17).value=BorderNum
            PassportNum = frame.getPassNum(Nationality)
            sh.cell(Exlpnt, 18).value = PassportNum
            DoB=frame.getDob()
            print(DoB)
            sh.cell(Exlpnt,6).value=DoB
            #Table contents
            TblOccupation,Tblbscwage,Tblhousing,Tblcommision,Tblotherallwnce,Tbltotalwage=frame.gettabledata()
            sh.cell(Exlpnt,7).value=TblOccupation
            sh.cell(Exlpnt,8).value=Tblbscwage
            sh.cell(Exlpnt,9).value=Tblhousing
            sh.cell(Exlpnt,10).value=Tblcommision
            sh.cell(Exlpnt,11).value=Tblotherallwnce
            sh.cell(Exlpnt,12).value=Tbltotalwage
            EmpId = frame.getEmpID()
            sh.cell(Exlpnt,13).value=EmpId
            EndContrReason =frame.getEndreason(Status)
            Cntrctstatus =frame.getContrStatus() # depends on the nationality
            sh.cell(Exlpnt,14).value=Cntrctstatus
            Engagementstatus=frame.getEngagementstatus()
            sh.cell(Exlpnt,15).value=Engagementstatus
            wb.save(path)
            frame.back()
            time.sleep(1)
          except Exception as s:
            wb.save(path)
            print('The error was ',s)
            text = "The error was" + str(s) + '\n'
            wb.save(path)
            frame.back()
            time.sleep(1)
            continue

end = time.time()
hours, rem = divmod(end-start, 3600)
minutes, seconds = divmod(rem, 60)
print("{:0>2}:{:0>2}:{:05.2f}".format(int(hours),int(minutes),seconds))
wb.save(path)
#frame.close()






